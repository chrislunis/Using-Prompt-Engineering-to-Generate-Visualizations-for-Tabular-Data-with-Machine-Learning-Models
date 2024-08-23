from xml.etree.ElementInclude import include
import streamlit as st
import pandas as pd
from api import OpenAI_API, open_ai_response
from question_bank import question_bank
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl

def page_config():
    st.set_page_config(page_title="Prompt Data Analyis and Visualization", layout="centered")
    hide_menu_style = "<style> footer {visibility: hidden;} </style>"
    st.markdown(hide_menu_style, unsafe_allow_html=True)

def display_relevant_plots(df):
    # Example plot: Distribution of a numeric column
    st.subheader("Distribution of a Numeric Column")
    fig, ax = plt.subplots()
    sns.histplot(df.select_dtypes(include=['float64', 'int64']).iloc[:, 0], kde=True, ax=ax)
    st.pyplot(fig)
    
    # Box plot for a categorical column vs. numeric column
    if df.select_dtypes(include=['object']).shape[1] > 0 and df.select_dtypes(include=['float64', 'int64']).shape[1] > 0:
        st.subheader("Box Plot of Categorical Column vs. Numeric Column")
        fig, ax = plt.subplots()
        sns.boxplot(x=df.select_dtypes(include=['object']).columns[0], y=df.select_dtypes(include=['float64', 'int64']).columns[0], data=df, ax=ax)
        st.pyplot(fig)
    
    # Correlation heatmap
    st.subheader("Correlation Heatmap")
    fig, ax = plt.subplots(figsize=(12, 8))
    sns.heatmap(df.corr(), annot=True, cmap='coolwarm', ax=ax)
    st.pyplot(fig)

def check_formatting(upload_file):
    wb = openpyxl.load_workbook(upload_file)
    formatting_info = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        conditional_formats = []
        charts = []
        formulas = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.has_style and cell.style.alignment is not None:
                    conditional_formats.append((cell.coordinate, cell.style.alignment))
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append(cell.coordinate)
        
        for chart in sheet._charts:
            charts.append(chart.title)
        
        formatting_info[sheet_name] = {
            "conditional_formats": conditional_formats,
            "charts": charts,
            "formulas": formulas
        }
    
    return formatting_info

def run_app():
    api_key = st.secrets["api_key"]['open_ai']
    if api_key is None:
        st.error('API key not found. Please set the api_key in the .streamlit/secrets.toml file.')
    else:
        with OpenAI_API(api_key):
            st.title('Prompt Data Analysis and Visualization')

            upload_file = st.file_uploader("Upload your data file (CSV or Excel)", type=['csv', 'xlsx', 'xls'])

            if upload_file is not None:
                if upload_file.name.endswith('.csv'):
                    df = pd.read_csv(upload_file)
                    dfs = {"Sheet1": df}
                    sheet_names = list(dfs.keys())
                    formatting_info = {}
                elif upload_file.name.endswith('.xlsx') or upload_file.name.endswith('.xls'):
                    xls = pd.ExcelFile(upload_file)
                    dfs = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}
                    sheet_names = xls.sheet_names
                    formatting_info = check_formatting(upload_file)

                selected_sheet = st.selectbox('Select a sheet to analyze', sheet_names)
                df = dfs[selected_sheet]
                prompt_selected = st.selectbox('Select a prompt', list(question_bank.keys()))
                prompt = question_bank[prompt_selected]['question']
                st.write(prompt)
                submit = st.button('Generate response')

                if submit:
                    with st.spinner('Generating response...'):
                        if prompt_selected == "DisplayRelevantPlots":
                            display_relevant_plots(df)
                        else:
                            answer = open_ai_response(prompt, df, selected_sheet)
                            st.markdown('### Output:')
                            st.write(answer)
                            
                            if formatting_info:
                                st.markdown('### Formatting Information:')
                                st.write(formatting_info[selected_sheet])
            else:
                st.error('Please upload a data file')
                st.stop()

if __name__ == '__main__':
    page_config()
    run_app()
